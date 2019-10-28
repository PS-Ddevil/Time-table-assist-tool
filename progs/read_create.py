import pandas as pd
import ezodf
import pyexcel
import os
import xlsxwriter
import openpyxl
from tkinter import filedialog
file_name = filedialog.askdirectory()

file_name = str(os.path.join(file_name,"CourselistAug-Dec2019_2Sep.ods"))


sheet =pyexcel.get_sheet(file_name=file_name)

for row in sheet:
    for p in range(len(row)):
        if(row[p]=='C'):    
            if(os.path.exists(str(p) + ".xlsx")):
                wb = openpyxl.load_workbook(str(p) + ".xlsx")
                pass
            else:
                wb = openpyxl.Workbook()
                wb.save(str(p) + ".xlsx")
                wb = openpyxl.load_workbook(str(p) + ".xlsx")
            # print(df_dict)
            ws = wb.worksheets[0]
            #headers = df_dict.keys()
            #lists=[]
            #lists=list(headers)
            
            #print(df_dict)
            #print(len(lists),p)
            rows = ws.max_row+1
            for l in range(6):
                ws.cell(rows,l+1).value=row[l]
                wb.save(str(p) + ".xlsx")
            ws.cell(rows,6).value=row[p]


                
            # k=0
            # rows = ws.max_row
            # # print("file name ----",str(p),"   ",rows)
            # for(idx,header) in enumerate(headers,start=1):
            #     # print("----------",idx,header,"-----------")
            #     ws.cell(row=rows+1,column= idx,value=header)
            #     wb.save(str(p) + ".xlsx")
            wb.save(str(p) + ".xlsx")





        






