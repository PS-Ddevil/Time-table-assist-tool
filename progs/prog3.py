from tkinter import *
from tkinter import messagebox
import os
import openpyxl
from tkinter import filedialog
from tkscrolledframe import ScrolledFrame



file_path = "src/tmp/baskets"
lists=[]
master2 = Tk()
master2.title("Class Selection")
master2.geometry("500x500+300+300")
frame1=Frame(master2,bd=10,height=500,padx=10,pady=10,width=1500)
v = StringVar(frame1)
frame1.pack()
slotA=[]
slotB=[]
slotC=[]
slotD=[]
slotE=[]
slotF=[]
slotG=[]
slotH=[]
sf = ScrolledFrame(master2, width=640, height=480)
sf.pack(side="top", expand=1, fill="both")

# Bind the arrow keys and scroll wheel
sf.bind_arrow_keys(master2)
sf.bind_scroll_wheel(master2)

# Create a frame within the ScrolledFrame
frame2 = sf.display_widget(Frame)

classes=['No Classroom','A1-3','A1-NKN','A5-1','A5-2','A5-4','A5-5','A9-1','SC-NKN','G-106,107','A10-1a','A10-1b','A10-1c','A10-1d','A10-2a','A10-2b','A10-2c','A10-3a','A10-3b','A10-3c','Hall A','Hall B','Hall C','A13-1A','A13-3A','A13-2A','A13-2B','A13-2C','A13-2D']
def ConstraintCheck(var,i,name,slot,*args):
    for j in range(len(slot)):
        if(i==j):
            continue
        elif(var.get()!="No Classroom" and var.get()==slot[j][1]):
            messagebox.showerror("Error", var.get() + " is twice time and it will not be saved so please change ") 
            var.set(slot[i][1])
            return
    slot[i][1]=var.get()
def Save():
    for file in lists:
        a=str(os.path.join(file_path, file))
        wb = openpyxl.load_workbook(a)
        ws = wb.active
        ws.cell(1,9).value="Classroom"
        for i in range(2,ws.max_row+1):
            if(ws.cell(i,8).value=="Slot A"):
                for j in range(len(slotA)):
                    if(slotA[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotA[j][1]
            elif(ws.cell(i,8).value=="Slot B"):
                for j in range(len(slotB)):
                    if(slotB[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotB[j][1]
            elif(ws.cell(i,8).value=="Slot C"):
                for j in range(len(slotC)):
                    if(slotC[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotC[j][1]
            elif(ws.cell(i,8).value=="Slot D"):
                for j in range(len(slotD)):
                    if(slotD[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotD[j][1]
            elif(ws.cell(i,8).value=="Slot E"):
                for j in range(len(slotE)):
                    if(slotE[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotE[j][1]
            elif(ws.cell(i,8).value=="Slot F"):
                for j in range(len(slotF)):
                    if(slotF[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotF[j][1]
            elif(ws.cell(i,8).value=="Slot G"):
                for j in range(len(slotG)):
                    if(slotG[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotG[j][1]
            else:
                for j in range(len(slotH)):
                    if(slotH[j][0]==ws.cell(i,1).value):
                        ws.cell(i,9).value=slotH[j][1] 
        wb.save(os.path.join(file_path, file))   

def call(slot):
    w = Label(frame2, text="Course")
    w.grid(row=0,column=0,columnspan=1,sticky=W+E+N+S)
    w.config(width=30)
    w = Label(frame2, text="Classroom")
    w.grid(row=0,column=1,columnspan=1,sticky=W+E+N+S)
    w.config(width=30)
    i=0
    for i in range(len(slot)):
        print(slot[i],"      fdfddfvfdv")
        w = Label(frame2, text=slot[i][0])
        w.grid(row=i+1,column=0,columnspan=1,sticky=W+E+N+S)
        w.config(width=30)
        var=StringVar()
        var.set(slot[i][1])
        w = OptionMenu(frame2,var,*classes)
        w.grid(row=i+1,column=1,sticky=W+E+N+S)
        w.config(width=30)
        var.trace("w", lambda name1,name2,op,i=i,var=var,name=v.get(),slot=slot:ConstraintCheck(var,i,name,slot,name1,name2,op))
        
    w = Button(frame2, text="Save",command=Save)
    w.grid(row=i+3,column=0,columnspan=2,sticky=W+E+N+S)
    w.config(width=15)
def DisplayGUI(*args):
    global file_path
    for widget in frame2.winfo_children():
        widget.destroy()
    if(v.get()=="Slot A"):
        call(slotA)
    elif(v.get()=="Slot B"):
        call(slotB)
    elif(v.get()=="Slot C"):
        call(slotC)
    elif(v.get()=="Slot D"):
        call(slotD)
    elif(v.get()=="Slot E"):
        call(slotE)
    elif(v.get()=="Slot F"):
        call(slotF)
    elif(v.get()=="Slot G"):
        call(slotG)
    else:
        call(slotH)    

    
def DropDown():
    global lists,v,slotA,slotB,slotC,slotD,slotE,slotF,slotG
    v.set("Select Slot")
    for file in lists:
        a=str(os.path.join(file_path, file))
        wb = openpyxl.load_workbook(a)
        ws = wb.active
        for i in range(2,ws.max_row+1):
            if(ws.cell(i,8).value=="Slot A"):
                if(ws.cell(i,9).value is None):
                    if [ws.cell(i,1).value,"No Classroom"] not in slotA:
                        slotA.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotA:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                    
            elif(ws.cell(i,8).value=="Slot B"):
                if(ws.cell(i,9).value is None):
                    if [ws.cell(i,1).value,"No Classroom"] not in slotB:
                        slotB.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotB:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                    
            elif(ws.cell(i,8).value=="Slot C"):
                if(ws.cell(i,9).value is None):
                    if [ws.cell(i,1).value,"No Classroom"] not in slotC:
                        slotC.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotC:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                    
            elif(ws.cell(i,8).value=="Slot D"):
                if(ws.cell(i,9).value is None):
                    if [ws.cell(i,1).value,"No Classroom"] not in slotD:
                        slotD.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotD:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                    
            elif(ws.cell(i,8).value=="Slot E"):
                if(ws.cell(i,9).value is None):
                    if [ws.cell(i,1).value,"No Classroom"] not in slotE:
                        slotE.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotE:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                    
            elif(ws.cell(i,8).value=="Slot F"):
                if(ws.cell(i,9).value is None):
                    if [ws.cell(i,1).value,"No Classroom"] not in slotF:
                        slotF.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotF:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                    
            elif(ws.cell(i,8).value=="Slot G"):
                if(ws.cell(i,9).value is None):
                        if [ws.cell(i,1).value,"No Classroom"] not in slotG:
                            slotG.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotG:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                    
            elif(ws.cell(i,8).value=="Slot H"):
                if(ws.cell(i,9).value is None):
                    if [ws.cell(i,1).value,"No Classroom"] not in slotH:
                        slotH.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    if [ws.cell(i,1).value,"No Classroom"] not in slotH:
                        slotA.append([ws.cell(i,1).value,ws.cell(i,9).value])                        


    slots=['Slot A','Slot B','Slot C','Slot D','Slot E','Slot F','Slot G','Slot H']
    
    print(slotA)
    print("    b jsd jfh dsfhdsf   ")
    print(slotB)
    print("    b jsd jfh dsfhdsf   ")
    print(slotC)
    print("    b jsd jfh dsfhdsf   ")
    print(slotD)
    print("    b jsd jfh dsfhdsf   ")
    print(slotE)
    print("    b jsd jfh dsfhdsf   ")
    print(slotF)
    print("    b jsd jfh dsfhdsf   ")
    print(slotG)
    print("    b jsd jfh dsfhdsf   ")
    print(slotH)
    print("    b jsd jfh dsfhdsf   ")
    w = OptionMenu(frame1, v, *slots)
    w.grid(row=1,column=0,columnspan=3,sticky=W+E+N+S)
    w.config(width=45)
    v.trace("w", DisplayGUI)

def DialogBox():
       global file_path
       
       for file in os.listdir(file_path):
           if file.endswith('.xlsx') and file!=".xlsx"  and file!="course_faculty_main.xlsx"  and file!='course_faculty_optional.xlsx':
               a=str(os.path.join(file_path, file))
               lists.append(file)
       DropDown()

B = Button(frame1, text ="Start", command = DialogBox)
B.grid(row=0,column=0,columnspan=3,sticky=W+E+N+S)
B.config(width=45)

mainloop()