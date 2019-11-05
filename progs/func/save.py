from openpyxl import load_workbook
import xlwt
def save_in_excel_sheet(array):
    fname='/home/vinay/Documents/Learn/Time-table-assist-tool/src/data/slot.xlsx'
    workbook=load_workbook(fname)
    sheet = workbook.active
    for i in range(len(array)):
        if (i%9)+2!=6:
            sheet.cell(row = int(i/9)+2, column = (i%9)+2).value=array[i].get()
    workbook.save("/home/vinay/Documents/Learn/Time-table-assist-tool/src/data/slot.xlsx")
    print('\nSAVED SUCCESSFULLY !\n')
