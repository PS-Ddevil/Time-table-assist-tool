import pandas as pd
import ezodf
import pyexcel
import os
import xlsxwriter
import openpyxl



file_path = "src/tmp/baskets"
def read_create():
    wb = openpyxl.Workbook()
    wb.save("src/tmp/baskets/"+"left_over" + ".xlsx")
    
    file_name = "src/data/Courselist.ods"

    sheet =pyexcel.get_sheet(file_name=file_name)
    name=[]
    i=0
    for row in sheet:
        lists=str(row[3])
        lists=lists.split('-')
        if(lists[0]!="0" ):
            check=0
            i+=1
            if(i==2):
                continue
            if(i>186):
                break
            # print(row)
            # print("sdjhfjdsgf hjdshf hds              ddvfdvhghjd d f                        ",i)

            for file in os.listdir(file_path):
                if file.endswith('.xlsx') and file!=".xlsx" and file!="course_faculty_main.xlsx"  and file!='course_faculty_optional.xlsx':
                    
                    a=str(os.path.join(file_path, file))
                    if(check==1 or row[1] != ""):
                        break
                    wb = openpyxl.load_workbook(a)
                    ws = wb.worksheets[0]
                    j=0
                    for j in range(2,ws.max_row+1):
                        # print(row[1],ws.cell(j,1).value)
                        if(ws.cell(j,1).value==row[1]):
                            check=1
                            break
                    wb.save(a)
            if(check==0):
                wb = openpyxl.load_workbook("src/tmp/baskets/"+"left_over" + ".xlsx")
                ws = wb.worksheets[0]
                rows = ws.max_row+1
                # print(rows)
                for l in range(1,6):
                    ws.cell(rows,l).value=row[l]

                    wb.save("src/tmp/baskets/"+"left_over" + ".xlsx")
    
    

read_create()