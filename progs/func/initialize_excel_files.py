import xlsxwriter
import xlwt
import os.path
import openpyxl

def slot():
    fname='src/tmp/slot.xlsx'
    if os.path.isfile(fname):
        print('slot.xlsx is already there')
    else:
        print('slot.xlsx is created')
        workbook=xlwt.Workbook(fname)
        ws = workbook.add_sheet('Tested')
        workbook.save(fname)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row = 1, column = 1+1).value='8:00 - 8:50'
        sheet.cell(row = 1, column = 2+1).value='9:00 - 9:50'
        sheet.cell(row = 1, column = 3+1).value='10:00 - 10:50'
        sheet.cell(row = 1, column = 4+1).value='11:00 - 11:50'
        # sheet.cell(row = 1, column = 5+1).value='12:00 - 12:50'
        sheet.cell(row = 1, column = 6).value='LUNCH BREAK'
        sheet.cell(row = 1, column = 7).value='1:00 - 1:50'
        sheet.cell(row = 1, column = 8).value='2:00 - 2:50'
        sheet.cell(row = 1, column = 9).value='3:00 - 3:50'
        sheet.cell(row = 1, column = 10).value='4:00 - 4:50'

        sheet.cell(row = 2, column = 1).value='Monday'
        sheet.cell(row = 3, column = 1).value='Tuesday'
        sheet.cell(row = 4, column = 1).value='Wednesday'
        sheet.cell(row = 5, column = 1).value='Thrusday'
        sheet.cell(row = 6, column = 1).value='Friday'
        wb.save("src/tmp/slot.xlsx")
