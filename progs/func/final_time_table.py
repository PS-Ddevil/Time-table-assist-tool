from tkinter import *
from tkinter import messagebox
import os
import openpyxl
from tkinter import filedialog
import tkinter.ttk as ttk


lists=[]
matrix = [[] for _ in range(11)]
mat = [[] for _ in range(30)]
file_path = "./src/tmp/baskets"

slotA=[]
slotB=[]
slotC=[]
slotD=[]
slotE=[]
slotF=[]
slotG=[]
slotH=[]
def final_time_table():
    global slotA,slotB,slotC,slotD,slotE,slotF,slotG,slotH
        # print(matrix[j])
    for file in os.listdir(file_path):
        if file.endswith(".xlsx") and file != ".xlsx" and file!="course_faculty_main.xlsx" and file!='course_faculty_optional.xlsx':

            lists.append(file)
            # print(file)
            a=str(os.path.join(file_path, file))
            wb = openpyxl.load_workbook(a)
            ws = wb.active
            # print(a)
            for i in range(1,ws.max_row+1):
                # print(ws.cell(i,8).value)
                if(str(ws.cell(i,8).value) == "Slot A"):
                    slotA.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
                elif(str(ws.cell(i,8).value) == "Slot B"):
                    slotB.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
                elif(str(ws.cell(i,8).value) == "Slot C"):
                    slotC.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
                elif(str(ws.cell(i,8).value) == "Slot D"):
                    slotD.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
                elif(str(ws.cell(i,8).value) == "Slot E"):
                    slotE.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
                elif(str(ws.cell(i,8).value) == "Slot F"):
                    slotF.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
                elif(str(ws.cell(i,8).value) == "Slot G"):
                    slotG.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
                elif(str(ws.cell(i,8).value) == "Slot H"):
                    slotH.append(str(ws.cell(i,1).value) +" - (" + str(ws.cell(i,9).value) + ")")
    
    

    wb = openpyxl.Workbook()
    wb.save("Courses-Slot-wise" + ".xlsx")
    wb = openpyxl.load_workbook("Courses-Slot-wise" + ".xlsx")
    ws = wb.worksheets[0]
    # for i in range(0,10):
    #     matrix[i]=list(dict.fromkeys(matrix[i]))
    
    # for i in range(10):
    #     for j in range(0,len(matrix[i])):
    #         mat[j].append(matrix[i][j])
    # #     print("----------------")
    # #     print(matrix[i])
    # #     print("----------------")
    # print(mat)


    
    ws.cell(1,1).value="Slot A"
    slotA=list(set(slotA))
    for i in range(len(slotA)):
        ws.cell(i+2,1).value=slotA[i]
    ws.cell(1,2).value="Slot B"
    slotB=list(set(slotB))
    for i in range(len(slotB)):
        ws.cell(i+2,2).value=slotB[i]
    ws.cell(1,3).value="Slot C"
    slotC=list(set(slotC))
    for i in range(len(slotC)):
        ws.cell(i+2,3).value=slotC[i]
    ws.cell(1,4).value="Slot D"
    slotD=list(set(slotD))
    for i in range(len(slotD)):
        ws.cell(i+2,4).value=slotD[i]
    ws.cell(1,5).value="Slot E"
    slotE=list(set(slotE))
    for i in range(len(slotE)):
        ws.cell(i+2,5).value=slotE[i]
    ws.cell(1,6).value="Slot F"
    slotF=list(set(slotF))
    for i in range(len(slotF)):
        ws.cell(i+2,6).value=slotF[i]
    ws.cell(1,7).value="Slot G"
    slotG=list(set(slotG))
    for i in range(len(slotG)):
        ws.cell(i+2,7).value=slotG[i]
    ws.cell(1,8).value="Slot H"
    slotH=list(set(slotH))
    for i in range(len(slotH)):
        ws.cell(i+2,8).value=slotH[i]


    wb.save("Courses-Slot-wise" + ".xlsx")

final_time_table()   