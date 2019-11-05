import pandas as pd
import ezodf
# import pandas as pd
# from docx.api import Document
from openpyxl import load_workbook
import xlwt
import openpyxl
import os
def extract():
    # pass
    fname1='./src/tmp/baskets/course_faculty_main.xlsx'
    print(str(os.path.isfile(fname1)))
    if str(os.path.isfile(fname1))=='False':
        print(os.path.isfile(fname1))
        workbook1=xlwt.Workbook(fname1)
        ws1 = workbook1.add_sheet('Tested')
        workbook1.save(fname1)
        print(fname1,' is created')

    workbook1=xlwt.Workbook(fname1)
    # workbook1.save(fname)
    wb1 = openpyxl.Workbook()
    sheet1 = wb1.active
    sheet1.cell(row = 1, column = 1).value='Course code'
    sheet1.cell(row = 1, column = 2).value='Faculty name'
    ##################################################################
    fname2='./src/tmp/baskets/course_faculty_optional.xlsx'
    print(str(os.path.isfile(fname2)))
    if str(os.path.isfile(fname2))=='False':
        print(os.path.isfile(fname2))
        workbook2=xlwt.Workbook(fname2)
        ws2 = workbook2.add_sheet('Tested')
        workbook2.save(fname2)
        print(fname2,' is created')

    workbook2=xlwt.Workbook(fname2)
    # workbook1.save(fname)
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.cell(row = 1, column = 1).value='Course code'
    sheet2.cell(row = 1, column = 2).value='Faculty name'
    #####################################################################
    count1=2
    count2=2

    doc = ezodf.opendoc('./src/data/Courselist.ods')

    sheet = doc.sheets[0]
    print(sheet)
    df_dict = {}
    check=False
    for i, row in enumerate(sheet.rows()):
        if i >= 2:
            df_dict = {cell.value:[] for cell in row}
            col_index = {j:cell.value for j, cell in enumerate(row)}
            if (type((col_index[0])) is float)==False:
                check=True
            else:
                faculty_name=str(col_index[4]).split(',')
                
                sheet1.cell(row = count1, column = 1).value=col_index[1]
                sheet1.cell(row = count1, column = 2).value=faculty_name[0]
                count1=count1+1

                if len(faculty_name)>1:
                    sheet2.cell(row = count2, column = 1).value=col_index[1]
                    sheet2.cell(row = count2, column = 2).value=col_index[4]
                    count2=count2+1

        if check==True:
            break
    # and convert to a DataFrame
    # df = pd.DataFrame(df_dict)
    wb1.save(fname1)
    wb2.save(fname2)

extract()
