from tkinter import *
from tkinter import messagebox
import os
import openpyxl
from tkinter import filedialog
from tkscrolledframe import ScrolledFrame

print("running-----------------------------------------------------------------")
file_path ='./src/tmp/baskets/course_faculty_optional.xlsx'
master2 = Tk()
master2.title("Multi Faculty List")
master2.geometry("750x500+300+300")

sf = ScrolledFrame(master2, width=640, height=400)
sf.pack(side="top", expand=1, fill="both")

# Bind the arrow keys and scroll wheel
sf.bind_arrow_keys(master2)
sf.bind_scroll_wheel(master2)

# Create a frame within the ScrolledFrame
frame2 = sf.display_widget(Frame)

wb = openpyxl.load_workbook(file_path)
ws = wb.active
w = Label(frame2, text="Course",borderwidth=2)
w.grid(row=0,column=0,columnspan=1,sticky=W+E+N+S)
w.config(width=30)
w = Label(frame2, text="Multiple Faculty",borderwidth=2)
w.grid(row=0,column=1,columnspan=2,sticky=W+E+N+S)
w.config(width=30)
for i in range(2,ws.max_row+1):
    w = Label(frame2, text=ws.cell(i,1).value,borderwidth=2)
    w.grid(row=i-1,column=0,columnspan=1,sticky=W+E+N+S)
    w.config(width=30)
    w = Label(frame2, text=ws.cell(i,2).value,borderwidth=2)
    w.grid(row=i-1,column=1,columnspan=2,sticky=W+E+N+S)
    w.config(width=60)   
wb.save(file_path)
frame=Frame(master2,bd=10,height=500,padx=10,pady=10,width=1500)
frame.pack()
w = Label(frame, fg="RED",text="*Note:- The followiing courses multiple professor alloted, kindly go to the following location ")
w.pack()
w = Label(frame, fg="RED",text="'/Time-table-assist-tool/src/tmp/baskets/course_faculty_optional.xlsx' and do the following changes to it.")
w.pack()

print("")
mainloop() 
