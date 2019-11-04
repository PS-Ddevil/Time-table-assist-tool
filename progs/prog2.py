# Include your packages here
from tkinter import *
from tkinter import messagebox
import os
import openpyxl
from tkinter import filedialog
import tkinter.ttk as ttk
from tkscrolledframe import ScrolledFrame

file_path = "src/tmp/baskets"
master2 = Tk()
master2.title("SLot Selection")
master2.geometry("500x500+100+100")
frame1=Frame(master2,bd=10,height=500,padx=10,pady=10,width=1500)
frame1.pack()

v = StringVar(frame1)
lists=[]
def DropDown():
    global lists,v
    v.set("Select File")
    w = ttk.Combobox(frame1, textvariable=v, value=lists)
    w.grid(row=1,column=0,columnspan=3,sticky=W+E+N+S)
    w.config(width=45)
    v.trace("w", DisplayGUI)


vars = []
def ConstraintCheck(var,i,name,focus,*args):
    global file_path
    wb = openpyxl.load_workbook(str(os.path.join(file_path, v.get())))
    ws = wb.active
    print(ws.cell(i+1,8).value,ws.cell(i,8).value,i,v.get(),file_path)
    j=0
    for j in range(2,ws.max_row+1):
        c1=ws.cell(j,8)
        if i==j:
            continue
        if c1.value==var.get():
            messagebox.showerror("Error", var.get() + " is twice time and it will not be saved so please change ") 
            var.set(ws.cell(i,8).value)
            return
    wb.save(str(os.path.join(file_path, v.get())))
    prof=ws.cell(i,7).value
    slot=ws.cell(i,8).value
    course=ws.cell(i,1).value
    for file in lists:
        if(file==v.get()):
            continue
        wb = openpyxl.load_workbook(str(os.path.join(file_path, file)))
        ws=wb.active
        for k in range(2,ws.max_row+1):
            if ws.cell(k,7).value==prof and ws.cell(k,8).value==var.get() and ws.cell(k,1).value!=course:
                
                messagebox.showerror("Error", prof + " has same slot in "+ file)
                wb.save(str(os.path.join(file_path, file)))
                var.set(slot)
                
                return
        wb.save(str(os.path.join(file_path, file)))
       
    wb = openpyxl.load_workbook(str(os.path.join(file_path, v.get())))
    ws = wb.active
    
    c1=ws.cell(i,8)
    c1.value=var.get()
    
        
    wb.save(str(os.path.join(file_path, v.get())))

frame2=Frame(master2,bd=10,height=500,padx=10,pady=10,width=1500)
frame2.pack()

def DisplayGUI(*args):
    global file_path
    for widget in frame2.winfo_children():
        widget.destroy()
    wb = openpyxl.load_workbook(str(os.path.join(file_path, v.get())))
    ws = wb.active
    slots=['Slot A','Slot B','Slot C','Slot D','Slot E','Slot F','Slot G','Slot H']
    c1=ws.cell(1,8)
    if c1.value is None: 
        c1.value="Slots"
    w = Label(frame2, text="Course")
    w.grid(row=2,column=0,columnspan=1,sticky=W+E+N+S)
    w = Label(frame2, text="Faculty")
    w.grid(row=2,column=1,columnspan=1,sticky=W+E+N+S)
    w = Label(frame2, text="Slots")
    w.grid(row=2,column=2,columnspan=1,sticky=W+E+N+S)
    for i in range(2,ws.max_row+1):
        c1=ws.cell(i,8)
        var = StringVar(frame2)
        if c1.value is None: 
            c1.value="No Slots"
            var.set("No Slots")
        else:
            var.set(c1.value)
        vars.append(var)
        w = Label(frame2, text=ws.cell(i,1).value)
        w.grid(row=i+2,column=0,columnspan=1,sticky=W+E+N+S)
        w.config(width=15)
        w = Label(frame2, text=ws.cell(i,7).value)
        w.grid(row=i+2,column=1,columnspan=1,sticky=W+E+N+S)
        w.config(width=15)
        w = OptionMenu(frame2,var,*slots)
        w.grid(row=i+2,column=2,sticky=W+E+N+S)
        w.config(width=15)
        var.trace("w", lambda name1,name2,op,i=i,var=var,name=v.get(),focus=w:ConstraintCheck(var,i,name,focus,name1,name2,op))
    wb.save(str(os.path.join(file_path, v.get())))

def DialogBox():
       global file_path
       
       global lists, v
       print("--------------------1")
       a=str(os.path.join(file_path, "course_faculty_main.xlsx"))
       print("--------------------2")
       wb = openpyxl.load_workbook(a)
       
       ws = wb.active
       course=[]
       faculty=[]
       for i in range(2,ws.max_row+1):
            course.append(ws.cell(i,1).value)
            faculty.append(ws.cell(i,2).value)
       wb.save(a)
       print(os.listdir(file_path))
       for file in os.listdir(file_path):
           if file.endswith('.xlsx') and file!=".xlsx" and file!="course_faculty_main.xlsx"  and file!='course_faculty_optional.xlsx':
               print(file)
               a=str(os.path.join(file_path, file))
               lists.append(file)
            #    print(a)
               wb = openpyxl.load_workbook(a)

               print("--------------------4")
               ws = wb.active
               if ws.cell(3,8).value is None :
                    ws.cell(1,7).value="Faculty name"
                    for i in range(2,ws.max_row+1):
                        c1=ws.cell(i,1).value
                        for j in range(len(course)):
                            if c1==course[j]:
                                ws.cell(i,7).value=faculty[j]
                                break
               wb.save(a)
       DropDown()
B = Button(frame1, text ="Start", command = DialogBox)
print("0000000000000000000000")
B.grid(row=0,column=0,columnspan=3,sticky=W+E+N+S)
B.config(width=45)

mainloop()