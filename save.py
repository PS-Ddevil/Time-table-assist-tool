from openpyxl import load_workbook
import xlwt
def save_in_excel_sheet(array):
    fname='Excel files/slot.xlsx'
    # print("i am in s,vldave inxxecel sheet ",array)
    workbook=load_workbook(fname)
    sheet = workbook.active
    for i in range(len(array)):
        if (i%9)+2!=6:
            sheet.cell(row = int(i/9)+2, column = (i%9)+2).value=array[i]
    workbook.save("Excel files/slot.xlsx")
