from tkinter import *
from tkinter import messagebox
import os
import openpyxl
from tkinter import filedialog
from tkinter import messagebox


master = Tk()
master.title("SLot Selection")
master.geometry("500x500+300+300")
#master.resizable(0, 0)
frame1=Frame(master,bd=10,height=500,padx=10,pady=10,width=1500)
frame1.pack()

v = StringVar(frame1)
lists=[]
def create():
    global lists,v
    print(lists)
    print(tuple(lists))
    v.set(lists[0])
    w = OptionMenu(frame1, v, *lists)
   # w= apply(OptionMenu, (frame1, v) + tuple(lists))
    w.grid(row=1,column=0,columnspan=3,sticky=W+E+N+S)
    w.config(width=45)
    v.trace("w", callback) 

vars = []
def callbacks(var,i,name,*args):
    
    print ("clicked at w00  ", var.get())   
    wb = openpyxl.load_workbook(v.get())
    ws = wb.active
    j=0
    for j in range(2,ws.max_row+1):
        c1=ws.cell(j,3)
        if i==j:
            continue
        if c1.value==var.get():
            messagebox.showerror("Error", var.get() + " is twice time and it will not be saved so please change ")
            break
    wb.save(v.get())
    prof=ws.cell(i,2).value
    print(ws.cell(3,3).value,"   fgffggfg  ",j,prof)
    for file in lists:
        if(file==v.get()):
            continue
        wb = openpyxl.load_workbook(file)
        ws=wb.active
        for k in range(2,ws.max_row+1):
            print(file,"  ", ws.cell(k,2).value,"  ",var.get()," ",ws.cell(k,3).value)
            if ws.cell(k,2).value==prof and ws.cell(k,3).value==var.get():
                
                messagebox.showerror("Error", prof + " has same slot in "+ file)
                wb.save(file)
                return
        wb.save(file)
       
    wb = openpyxl.load_workbook(v.get())
    ws = wb.active
    if j==ws.max_row :
        c1=ws.cell(i,3)
        c1.value=var.get()
    wb.save(v.get())


frame2=Frame(master,bd=10,height=500,padx=10,pady=10,width=1500)
frame2.pack()
def callback(*args):
    for widget in frame2.winfo_children():
        widget.destroy()
    print ("clicked at w00  ", v.get()  ) 
    wb = openpyxl.load_workbook(v.get())
    ws = wb.active
    slots=['Slot A','Slot B','Slot C','Slot D','Slot E','Slot F','Slot G','Slot H']
    c1=ws.cell(1,3)
    if c1.value is None: 
        c1.value="Slots"
    w = Label(frame2, text="Course")
    w.grid(row=2,column=0,columnspan=1,sticky=W+E+N+S)
    w = Label(frame2, text="Prof")
    w.grid(row=2,column=1,columnspan=1,sticky=W+E+N+S)
    w = Label(frame2, text="Slots")
    w.grid(row=2,column=2,columnspan=1,sticky=W+E+N+S)
    print(ws.max_row,"  jkkj  ")
    for i in range(2,ws.max_row+1):
        c1=ws.cell(i,3)
        var = StringVar(frame2)
        if c1.value is None: 
            c1.value="No Slots"
            var.set("No Slots")
        else:
            var.set(c1.value)
        vars.append(var)
        print(var)
        w = Label(frame2, text=ws.cell(i,1).value)
        w.grid(row=i+2,column=0,columnspan=1,sticky=W+E+N+S)
        w.config(width=15)
        w = Label(frame2, text=ws.cell(i,2).value)
        w.grid(row=i+2,column=1,columnspan=1,sticky=W+E+N+S)
        w.config(width=15)
        w = OptionMenu(frame2,var,*slots)
        #w= apply(OptionMenu, (frame2, var) + tuple(slots))
        w.grid(row=i+2,column=2,sticky=W+E+N+S)
        w.config(width=15)
        var.trace("w", lambda name1,name2,op,i=i,var=var,name=v.get():callbacks(var,i,name,name1,name2,op))
        print(c1.value)
    print(ws.max_row)
    print(ws.max_column)
    '''
    w = Label(frame1, text=ws.cell(2,2).value)
    w.grid(row=2,column=0,sticky=W+E+N+S)
    var = StringVar(frame1)
    var.set(slots[0])
    w= apply(OptionMenu, (frame1, var) + tuple(slots))
    w.grid(row=2,column=1,sticky=W+E+N+S)'''
    wb.save(v.get())
def helloCallBack():
       file_path=filedialog.askdirectory()
       print(file_path)
       global lists, v
       
       for file in os.listdir(file_path):
           if file.endswith('.xlsx'):
               #print(os.path.join(file_path, file))
               a=str(os.path.join(file_path, file))
               lists.append(a)
       create()
       #print(lists)
B = Button(frame1, text ="Choose Directory", command = helloCallBack)
B.grid(row=0,column=0,columnspan=3,sticky=W+E+N+S)
B.config(width=45)

mainloop()