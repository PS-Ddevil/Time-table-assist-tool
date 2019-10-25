from tkinter import *
from tkinter import messagebox
import os
import openpyxl
from tkinter import filedialog


file_path = ""
lists=[]
master2 = Tk()
master2.title("SLot Selection")
master2.geometry("500x500+300+300")
frame1=Frame(master2,bd=10,height=500,padx=10,pady=10,width=1500)
v = StringVar(frame1)
frame1.pack()
slotA=[]
slotB=[]
slotC=[]
slotD=[]
slotE=[]
slotF=[]
slotG=[]
slotH=[]
frame2=Frame(master2,bd=10,height=500,padx=10,pady=10,width=1500)
frame2.pack()
classes=['No Classroom','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15']
def ConstraintCheck(var,i,name,slot,*args):
    for j in range(len(slot)):
        if(i==j):
            continue
        elif(var.get()!="No Classroom" or var.get()==slot[j][1]):
            messagebox.showerror("Error", var.get() + " is twice time and it will not be saved so please change ") 
            var.set(slot[j][1])
            return
    slot[i][1]=var.get()
def Save():
    for file in lists:
        a=str(os.path.join(file_path, file))
        wb = openpyxl.load_workbook(a)
        ws = wb.active
        ws.cell(1,6).value="Classroom"
        for i in range(2,ws.max_row+1):
            if(ws.cell(i,5).value=="Slot A"):
                for j in range(len(slotA)):
                    if(slotA[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotA[j][1]
            elif(ws.cell(i,5).value=="Slot B"):
                for j in range(len(slotB)):
                    if(slotB[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotB[j][1]
            elif(ws.cell(i,5).value=="Slot C"):
                for j in range(len(slotC)):
                    if(slotC[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotC[j][1]
            elif(ws.cell(i,5).value=="Slot D"):
                for j in range(len(slotD)):
                    if(slotD[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotD[j][1]
            elif(ws.cell(i,5).value=="Slot E"):
                for j in range(len(slotE)):
                    if(slotE[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotE[j][1]
            elif(ws.cell(i,5).value=="Slot F"):
                for j in range(len(slotF)):
                    if(slotF[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotF[j][1]
            elif(ws.cell(i,5).value=="Slot G"):
                for j in range(len(slotG)):
                    if(slotG[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotG[j][1]
            else:
                for j in range(len(slotH)):
                    if(slotH[j][0]==ws.cell(i,1).value):
                        ws.cell(i,6).value=slotH[j][1] 
        ws.save(os.path.join(file_path, file))   

def call(slot):
    w = Label(frame2, text="Course")
    w.grid(row=0,column=0,columnspan=1,sticky=W+E+N+S)
    w.config(width=15)
    w = Label(frame2, text="Classroom")
    w.grid(row=0,column=0,columnspan=1,sticky=W+E+N+S)
    w.config(width=15)
    for i in range(len(slot)):
        w = Label(frame2, text=slot[i][0])
        w.grid(row=i+1,column=0,columnspan=1,sticky=W+E+N+S)
        w.config(width=15)
        var=StringVar()
        var.set(slot[i][1])
        w = OptionMenu(frame2,var,*classes)
        w.grid(row=i+1,column=1,sticky=W+E+N+S)
        w.config(width=15)
        var.trace("w", lambda name1,name2,op,i=i,var=var,name=v.get(),slot=slot:ConstraintCheck(var,i,name,slot,name1,name2,op))
        
    w = Button(frame2, text="Save",command=Save)
    w.grid(row=i,column=0,columnspan=2,sticky=W+E+N+S)
    w.config(width=15)
def DisplayGUI(*args):
    global file_path
    for widget in frame2.winfo_children():
        widget.destroy()
    if(v.get()=="Slot A"):
        call(slotA)
    elif(v.get()=="Slot B"):
        call(slotB)
    elif(v.get()=="Slot C"):
        call(slotC)
    elif(v.get()=="Slot D"):
        call(slotD)
    elif(v.get()=="Slot E"):
        call(slotE)
    elif(v.get()=="Slot F"):
        call(slotF)
    elif(v.get()=="Slot G"):
        call(slotG)
    else:
        call(slotH)    

    
def DropDown():
    global lists,v,slotA,slotB,slotC,slotD,slotE,slotF,slotG
    v.set("Select Slot")
    for file in lists:
        a=str(os.path.join(file_path, file))
        wb = openpyxl.load_workbook(a)
        ws = wb.active
        for i in range(2,ws.max_row+1):
            if(ws.cell(i,5).value=="Slot A"):
                if(ws.cell(i,6).value is None):
                    slotA.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                    
            elif(ws.cell(i,5).value=="Slot B"):
                if(ws.cell(i,6).value is None):
                    slotB.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                    
            elif(ws.cell(i,5).value=="Slot C"):
                if(ws.cell(i,6).value is None):
                    slotC.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                    
            elif(ws.cell(i,5).value=="Slot D"):
                if(ws.cell(i,6).value is None):
                    slotD.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                    
            elif(ws.cell(i,5).value=="Slot E"):
                if(ws.cell(i,6).value is None):
                    slotE.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                    
            elif(ws.cell(i,5).value=="Slot F"):
                if(ws.cell(i,6).value is None):
                    slotF.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                    
            elif(ws.cell(i,5).value=="Slot G"):
                if(ws.cell(i,6).value is None):
                        slotG.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                    
            else:
                if(ws.cell(i,6).value is None):
                    slotH.append([ws.cell(i,1).value,"No Classroom"])
                else:
                    slotA.append([ws.cell(i,1).value,ws.cell(i,6).value])                        


    slots=['Slot A','Slot B','Slot C','Slot D','Slot E','Slot F','Slot G','Slot H']
    print(slotA)
    print(slotB)
    print(slotC)
    print(slotD)
    print(slotE)
    print(slotF)
    print(slotG)
    print(slotH)
    w = OptionMenu(frame1, v, *slots)
    w.grid(row=1,column=0,columnspan=3,sticky=W+E+N+S)
    w.config(width=45)
    v.trace("w", DisplayGUI)

def DialogBox():
       global file_path
       file_path=filedialog.askdirectory()
       for file in os.listdir(file_path):
           if file.endswith('.xlsx') and file!="track.xlsx":
               a=str(os.path.join(file_path, file))
               lists.append(file)
       DropDown()

B = Button(frame1, text ="Choose Directory", command = DialogBox)
B.grid(row=0,column=0,columnspan=3,sticky=W+E+N+S)
B.config(width=45)

mainloop()